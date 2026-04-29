#!/usr/bin/env python3
"""
Generate repo_audit.xlsx from audit data.

Called automatically by audit.py — no need to run this separately.
Can also be run standalone against an existing audit_results.json:
    python audit/to_excel.py
(Note: audit.py no longer produces this JSON by default —
 the standalone path is kept for backward compatibility.)
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

RED = "FFC7CE"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
BLUE = "BDD7EE"
DARK = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"


def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def header_style(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = hfill(DARK)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_zebra(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        if r % 2 == 0:
            for c in range(1, cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF", "00FFFFFF"):
                    cell.fill = hfill("EEF3FA")


def autofit(ws, min_width=8, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


def freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def bool_str(v):
    return "Yes" if v else "No"


def trunc(s, n=80):
    s = str(s) if s else ""
    return s[:n] + ("..." if len(s) > n else "")


def highlight_bool_col(ws, col, start_row, end_row, yes_color=GREEN, no_color=RED):
    col_letter = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=hfill(yes_color))
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}",
        CellIsRule(operator="equal", formula=['"No"'], fill=hfill(no_color))
    )


def write_xlsx(repos, org_owners, audited_at, out_path):
    """Write the full audit workbook to out_path and return the path."""
    total = len(repos)
    org = repos[0].get("org", "") if repos else ""
    wb = Workbook()
    wb.remove(wb.active)

    # ── TAB 1: OVERVIEW ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("1. Overview")
    ws1.sheet_view.showGridLines = True
    h1 = [
        "Repo Name", "Visibility", "Archived?", "Fork?", "Language",
        "Description", "Has Description?", "Stars", "Open Issues",
        "Days Since Last Push", "Status (Active/Stale)",
        "Kebab-Case Name?", "Default Branch", "Homepage?",
    ]
    ws1.append(h1)
    header_style(ws1, 1, len(h1))
    for r in repos:
        ws1.append([
            r["name"], r["visibility"].capitalize(),
            bool_str(r["archived"]), bool_str(r["fork"]),
            r["language"], trunc(r["description"], 100),
            bool_str(r["has_description"]), r["stars"], r["open_issues"],
            r["last_push_days"] if isinstance(r["last_push_days"], int) else "?",
            "STALE" if r["stale"] else "Active",
            bool_str(r["kebab_case_name"]), r["default_branch"],
            bool_str(r["has_homepage"]),
        ])
    end1 = total + 1
    highlight_bool_col(ws1, 3, 2, end1, yes_color=YELLOW, no_color=GREEN)
    highlight_bool_col(ws1, 7, 2, end1)
    highlight_bool_col(ws1, 12, 2, end1)
    ws1.conditional_formatting.add(
        f"{get_column_letter(11)}2:{get_column_letter(11)}{end1}",
        CellIsRule(operator="equal", formula=['"STALE"'], fill=hfill(RED)))
    ws1.conditional_formatting.add(
        f"{get_column_letter(11)}2:{get_column_letter(11)}{end1}",
        CellIsRule(operator="equal", formula=['"Active"'], fill=hfill(GREEN)))
    ws1.conditional_formatting.add(
        f"{get_column_letter(2)}2:{get_column_letter(2)}{end1}",
        CellIsRule(operator="equal", formula=['"Private"'], fill=hfill(YELLOW)))
    ws1.conditional_formatting.add(
        f"{get_column_letter(2)}2:{get_column_letter(2)}{end1}",
        CellIsRule(operator="equal", formula=['"Public"'], fill=hfill(BLUE)))
    apply_zebra(ws1, 2, end1, len(h1))
    autofit(ws1)
    freeze(ws1, "B2")

    # ── TAB 2: DOCUMENTATION ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("2. Documentation")
    h2 = [
        "Repo Name", "Visibility",
        "README?", "README Chars", "README Lines", "README Quality", "README Placeholder Reason",
        "CITATION?", "Citation File", "CONTRIBUTING?", "Contributing File",
        "SECURITY Policy?", "Security File",
    ]
    ws2.append(h2)
    header_style(ws2, 1, len(h2))
    for r in repos:
        quality = "PLACEHOLDER" if r["readme_placeholder"] else "OK"
        ws2.append([
            r["name"], r["visibility"].capitalize(),
            bool_str(r["readme_exists"]), r["readme_chars"], r["readme_lines"],
            quality, r["readme_placeholder_reason"] if r["readme_placeholder"] else "--",
            bool_str(r["citation_exists"]), r["citation_path"],
            bool_str(r["contributing_exists"]), r["contributing_path"],
            bool_str(r["security_policy"]), r["security_path"],
        ])
    end2 = total + 1
    for ci in [3, 8, 10, 12]:
        highlight_bool_col(ws2, ci, 2, end2)
    ws2.conditional_formatting.add(
        f"{get_column_letter(6)}2:{get_column_letter(6)}{end2}",
        CellIsRule(operator="equal", formula=['"PLACEHOLDER"'], fill=hfill(YELLOW)))
    ws2.conditional_formatting.add(
        f"{get_column_letter(6)}2:{get_column_letter(6)}{end2}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=hfill(GREEN)))
    apply_zebra(ws2, 2, end2, len(h2))
    autofit(ws2)
    freeze(ws2, "B2")

    # ── TAB 3: LICENSE & TOPICS ───────────────────────────────────────────────
    ws3 = wb.create_sheet("3. License & Topics")
    h3 = ["Repo Name", "Visibility", "Has License?", "License SPDX", "License Name", "# Topics", "Topics / Tags"]
    ws3.append(h3)
    header_style(ws3, 1, len(h3))
    for r in repos:
        ws3.append([
            r["name"], r["visibility"].capitalize(),
            bool_str(r["license_spdx"] != "NONE"), r["license_spdx"], r["license_name"],
            r["topic_count"], r["topics"],
        ])
    end3 = total + 1
    highlight_bool_col(ws3, 3, 2, end3)
    ws3.conditional_formatting.add(
        f"F2:F{end3}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFC7CE",
                       end_type="num", end_value=10, end_color="C6EFCE"))
    apply_zebra(ws3, 2, end3, len(h3))
    autofit(ws3)
    freeze(ws3, "B2")

    # ── TAB 4: CONTRIBUTORS ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("4. Contributors")
    h4 = [
        "Repo Name", "Visibility", "# Contributors", "Last Committer", "Last Commit Date",
        "Contributor 1 (commits)", "Contributor 2", "Contributor 3", "Contributor 4", "Contributor 5",
    ]
    ws4.append(h4)
    header_style(ws4, 1, len(h4))
    for r in repos:
        top = r.get("top_contributors", [])

        def fmt(i):
            return f"{top[i][0]} ({top[i][1]})" if i < len(top) else "--"

        ws4.append([
            r["name"], r["visibility"].capitalize(), r["contributor_count"],
            r["last_contributor"], r["last_commit_date"],
            fmt(0), fmt(1), fmt(2), fmt(3), fmt(4),
        ])
    end4 = total + 1
    ws4.conditional_formatting.add(
        f"C2:C{end4}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFC7CE",
                       end_type="num", end_value=20, end_color="C6EFCE"))
    apply_zebra(ws4, 2, end4, len(h4))
    autofit(ws4)
    freeze(ws4, "B2")

    # ── TAB 5: GOVERNANCE & CI ────────────────────────────────────────────────
    ws5 = wb.create_sheet("5. Governance & CI")
    h5 = [
        "Repo Name", "Visibility", "CODEOWNERS?", "CODEOWNERS Path",
        "Branch Protection", "Has Protection?", "Rulesets Summary", "Has Rulesets?", "Any Protection?",
        "# CI Workflows", "Workflow Files",
        "PR Template?", "PR Template Path", "# Issue Templates", "Issue Template Names",
        "Dependabot?", "Vuln Alerts Enabled?",
    ]
    ws5.append(h5)
    header_style(ws5, 1, len(h5))
    for r in repos:
        has_prot = r["branch_protection"] not in ("none", "", None)
        ws5.append([
            r["name"], r["visibility"].capitalize(),
            bool_str(r["codeowners"]), r["codeowners_path"],
            r["branch_protection"], bool_str(has_prot),
            trunc(r["rulesets_summary"], 120), bool_str(r["has_rulesets"]), bool_str(r["has_any_protection"]),
            r["workflow_count"], trunc(r["workflows"], 100),
            bool_str(r["pr_template"]), r["pr_template_path"],
            r["issue_templates"], trunc(r["issue_template_names"], 80),
            bool_str(r["dependabot"]), bool_str(r["vuln_alerts"]),
        ])
    end5 = total + 1
    for ci in [3, 6, 8, 9, 12, 16, 17]:
        highlight_bool_col(ws5, ci, 2, end5)
    apply_zebra(ws5, 2, end5, len(h5))
    autofit(ws5)
    freeze(ws5, "B2")

    # ── TAB 6: SUMMARY SCORECARD ──────────────────────────────────────────────
    ws6 = wb.create_sheet("6. Summary Scorecard")
    ws6.append([f"GitHub Org Audit -- {org}", None, None, None, None])
    ws6["A1"].font = Font(bold=True, size=14, color=DARK)
    ws6.append([f"Audited: {audited_at}   |   Org Owners: {', '.join(org_owners)}", None, None, None, None])
    ws6.append([])
    ws6.append(["Check", "Pass Count", "Total", "Pass %", "Rating"])
    header_style(ws6, 4, 5)

    def rating(p):
        if p >= 80:
            return "Good"
        if p >= 50:
            return "Fair"
        if p >= 20:
            return "Poor"
        return "Critical"

    checks = [
        ("Has README",                   sum(1 for r in repos if r["readme_exists"])),
        ("Non-placeholder README",       sum(1 for r in repos if r["readme_exists"] and not r["readme_placeholder"])),
        ("Has License",                  sum(1 for r in repos if r["license_spdx"] != "NONE")),
        ("Has CONTRIBUTING",             sum(1 for r in repos if r["contributing_exists"])),
        ("Has SECURITY Policy",          sum(1 for r in repos if r["security_policy"])),
        ("Has CITATION",                 sum(1 for r in repos if r["citation_exists"])),
        ("Has Topics/Tags",              sum(1 for r in repos if r["topic_count"] > 0)),
        ("Has Description",              sum(1 for r in repos if r["has_description"])),
        ("Has CODEOWNERS",               sum(1 for r in repos if r["codeowners"])),
        ("Branch Protection Enabled",    sum(
            1 for r in repos
            if r["branch_protection"] not in ("none", "", None)
        )),
        ("Has CI Workflows",             sum(1 for r in repos if r["workflow_count"] > 0)),
        ("Has PR Template",              sum(1 for r in repos if r["pr_template"])),
        ("Has Issue Template(s)",        sum(1 for r in repos if r["issue_templates"] > 0)),
        ("Dependabot Enabled",           sum(1 for r in repos if r["dependabot"])),
        ("Vulnerability Alerts Enabled", sum(1 for r in repos if r["vuln_alerts"])),
        ("Kebab-Case Repo Name",         sum(1 for r in repos if r["kebab_case_name"])),
        ("Active (<=180 days)",          sum(1 for r in repos if not r["stale"])),
        ("Is Public",                    sum(1 for r in repos if r["visibility"] == "public")),
        ("Is Private",                   sum(1 for r in repos if r["visibility"] == "private")),
    ]
    score_start = 5
    for label, count in checks:
        p = round(100 * count / total) if total else 0
        ws6.append([label, count, total, p, rating(p)])
    score_end = score_start + len(checks) - 1

    for row in ws6.iter_rows(min_row=score_start, max_row=score_end, min_col=5, max_col=5):
        for cell in row:
            if cell.value == "Good":
                cell.fill = hfill(GREEN)
            elif cell.value == "Fair":
                cell.fill = hfill(YELLOW)
            elif cell.value == "Poor":
                cell.fill = hfill("F4B942")
            elif cell.value == "Critical":
                cell.fill = hfill(RED)

    ws6.conditional_formatting.add(
        f"D{score_start}:D{score_end}",
        ColorScaleRule(start_type="num", start_value=0, start_color=RED,
                       mid_type="num", mid_value=50, mid_color=YELLOW,
                       end_type="num", end_value=100, end_color="00C6EFCE"))
    ws6.append([])
    ws6.append(["Org Owners", ", ".join(org_owners)])
    ws6.cell(row=score_end + 3, column=1).font = Font(bold=True)
    ws6.append(["Note",
                "Branch protection requires admin auth; private repos may show none if token lacks write scope."])
    apply_zebra(ws6, score_start, score_end, 5)
    autofit(ws6, max_width=60)
    ws6.column_dimensions["A"].width = 35
    ws6.column_dimensions["D"].width = 10
    ws6.column_dimensions["E"].width = 12
    ws6.freeze_panes = "A5"

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    script_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(script_dir, "output", "audit_results.json")
    out_path = os.path.join(script_dir, "output", "repo_audit.xlsx")
    with open(json_path) as f:
        data = json.load(f)
    for r in data["repos"]:
        r.setdefault("org", data.get("org", ""))
    path = write_xlsx(data["repos"], data.get("org_owners", []), data.get("audited_at", ""), out_path)
    print(f"Saved: {path}")
